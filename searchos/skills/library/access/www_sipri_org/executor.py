"""SIPRI Military Expenditure and Arms Database Access Skill

This skill provides access to SIPRI's databases including:
- Military Expenditure Database (milex)
- Arms Industry Database
- Arms Transfers Database
- Publications and fact sheets
- Embedded Datawrapper charts

The skill downloads data directly from SIPRI's public file repository and
Datawrapper CDN without requiring browser automation.
"""

import asyncio
import csv
import io
import re
from typing import Any, Dict, List, Optional
import aiohttp
import openpyxl
from openpyxl.workbook import Workbook


class SIPRIAccessSkill:
    """Access skill for SIPRI databases and publications."""
    
    BASE_URL = "https://www.sipri.org"
    
    # Known database file URLs
    DATABASE_FILES = {
        "milex_full": {
            "url": "https://www.sipri.org/sites/default/files/SIPRI-Milex-data-1949-2025_v1.2.xlsx",
            "description": "Complete Military Expenditure Database (1949-2025)",
            "format": "xlsx",
            "category": "database"
        },
        "arms_industry_top100": {
            "url": "https://www.sipri.org/sites/default/files/SIPRI-Top-100-2002-2024%20%282%29.xlsx",
            "description": "SIPRI Top 100 Arms Industry Companies (2002-2024)",
            "format": "xlsx",
            "category": "database"
        },
        "arms_industry_totals": {
            "url": "https://www.sipri.org/sites/default/files/Total-arms-revenues-SIPRI-Top-100-2002-2024.xlsx",
            "description": "Total Arms Revenues for SIPRI Top 100 (2002-2024)",
            "format": "xlsx",
            "category": "database"
        }
    }
    
    # Known Datawrapper chart configurations
    DATAWRAPPER_CHARTS = {
        "milex_gdp_share": {
            "chart_id": "g7sno",
            "version": "12",
            "description": "Military expenditure as share of GDP by country",
            "category": "chart"
        }
    }
    
    def __init__(self):
        self.session: Optional[aiohttp.ClientSession] = None
        self.timeout = aiohttp.ClientTimeout(total=60)
    
    async def _get_session(self) -> aiohttp.ClientSession:
        """Get or create HTTP session."""
        if self.session is None or self.session.closed:
            self.session = aiohttp.ClientSession(timeout=self.timeout)
        return self.session
    
    async def _fetch_url(self, url: str) -> Dict[str, Any]:
        """Fetch content from URL."""
        session = await self._get_session()
        
        try:
            async with session.get(url) as response:
                if response.status != 200:
                    return {
                        "error": f"HTTP {response.status}",
                        "url": url
                    }
                
                content_type = response.headers.get('Content-Type', '')
                content = await response.read()
                
                return {
                    "content": content,
                    "content_type": content_type,
                    "url": url,
                    "size": len(content)
                }
        except asyncio.TimeoutError:
            return {"error": "Request timeout", "url": url}
        except Exception as e:
            return {"error": str(e), "url": url}
    
    async def _parse_xlsx(self, content: bytes) -> Dict[str, Any]:
        """Parse XLSX file and return structured data."""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            
            result = {
                "sheets": {},
                "sheet_names": workbook.sheetnames
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []
                
                # Get max rows and columns
                max_row = min(sheet.max_row, 1000)  # Limit to prevent huge outputs
                max_col = min(sheet.max_column, 50)
                
                for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                    row_data = []
                    for cell in row:
                        value = cell.value
                        if value is None:
                            row_data.append("")
                        elif isinstance(value, (int, float)):
                            row_data.append(value)
                        else:
                            row_data.append(str(value))
                    rows.append(row_data)
                
                result["sheets"][sheet_name] = {
                    "rows": rows,
                    "total_rows": sheet.max_row,
                    "total_columns": sheet.max_column,
                    "sample_rows": len(rows)
                }
            
            return result
        except Exception as e:
            return {"error": f"Failed to parse XLSX: {str(e)}"}
    
    async def _parse_csv(self, content: str) -> List[Dict[str, Any]]:
        """Parse CSV content into structured data."""
        try:
            reader = csv.DictReader(io.StringIO(content))
            return [row for row in reader]
        except Exception as e:
            # Try simple split if DictReader fails
            lines = content.strip().split('\n')
            if lines:
                header = lines[0].split(',')
                result = []
                for line in lines[1:]:
                    values = line.split(',')
                    if len(values) == len(header):
                        result.append(dict(zip(header, values)))
                return result
            return []
    
    async def download_database(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        Download and parse a SIPRI database file.
        
        Args:
            params: Must contain 'database' key with one of:
                    - "milex_full": Complete military expenditure database
                    - "arms_industry_top100": Top 100 arms companies
                    - "arms_industry_totals": Total arms revenues
        
        Returns:
            Parsed database data with sheets and rows
        """
        database_key = params.get("database", "")
        
        if database_key not in self.DATABASE_FILES:
            return {
                "error": f"Unknown database: {database_key}",
                "available_databases": list(self.DATABASE_FILES.keys())
            }
        
        db_info = self.DATABASE_FILES[database_key]
        
        # Fetch the file
        result = await self._fetch_url(db_info["url"])
        
        if "error" in result:
            return result
        
        # Parse based on file format
        if db_info["format"] == "xlsx":
            parsed = await self._parse_xlsx(result["content"])
            if "error" in parsed:
                return parsed
            
            return {
                "database": database_key,
                "description": db_info["description"],
                "url": db_info["url"],
                "size_bytes": result["size"],
                "data": parsed
            }
        else:
            return {
                "error": f"Unsupported format: {db_info['format']}"
            }
    
    async def get_datawrapper_data(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        Get data from embedded Datawrapper charts.
        
        Args:
            params: Can contain:
                    - "chart_id": Specific chart ID (e.g., "g7sno")
                    - "name": Chart name (e.g., "milex_gdp_share")
        
        Returns:
            Chart data as CSV rows
        """
        chart_id = params.get("chart_id")
        name = params.get("name")
        
        if name and name in self.DATAWRAPPER_CHARTS:
            chart_info = self.DATAWRAPPER_CHARTS[name]
            chart_id = chart_info["chart_id"]
            version = chart_info["version"]
        elif chart_id:
            version = params.get("version", "12")
        else:
            return {
                "error": "Must provide 'chart_id' or 'name'",
                "available_charts": list(self.DATAWRAPPER_CHARTS.keys())
            }
        
        # Construct CSV URL
        csv_url = f"https://datawrapper.dwcdn.net/{chart_id}/{version}/dataset.csv"
        
        # Fetch CSV
        result = await self._fetch_url(csv_url)
        
        if "error" in result:
            return result
        
        # Parse CSV
        csv_content = result["content"].decode('utf-8')
        parsed_data = await self._parse_csv(csv_content)
        
        return {
            "chart_id": chart_id,
            "version": version,
            "url": csv_url,
            "rows": len(parsed_data),
            "data": parsed_data[:100] if len(parsed_data) > 100 else parsed_data,
            "total_rows": len(parsed_data)
        }
    
    async def get_publication(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        Get information about a SIPRI publication and download if PDF.
        
        Args:
            params: Must contain 'url' or 'year' and 'type' and 'title', e.g.:
                    - {"url": "/publications/2025/sipri-fact-sheets/trends-world-military-expenditure-2024"}
                    - {"year": "2025", "type": "sipri-fact-sheets", "title": "trends-world-military-expenditure-2024"}
        
        Returns:
            Publication metadata and download links
        """
        url = params.get("url")
        
        if not url:
            year = params.get("year")
            pub_type = params.get("type")
            title = params.get("title")
            
            if not all([year, pub_type, title]):
                return {
                    "error": "Must provide 'url' or 'year', 'type', and 'title'"
                }
            
            url = f"/publications/{year}/{pub_type}/{title}"
        
        if not url.startswith("http"):
            url = f"{self.BASE_URL}{url}"
        
        # Fetch publication page
        result = await self._fetch_url(url)
        
        if "error" in result:
            return result
        
        html = result["content"].decode('utf-8')
        
        # Extract metadata
        metadata = {
            "url": url,
            "title": "",
            "pdf_files": [],
            "datawrapper_charts": []
        }
        
        # Extract title
        title_match = re.search(r'<title>([^<]+)</title>', html)
        if title_match:
            metadata["title"] = title_match.group(1).split('|')[0].strip()
        
        # Extract PDF download links
        pdf_links = re.findall(r'href="(/sites/default/files/[^"]+\.pdf)"', html)
        for pdf_link in pdf_links:
            pdf_url = f"{self.BASE_URL}{pdf_link}"
            filename = pdf_link.split('/')[-1]
            metadata["pdf_files"].append({
                "url": pdf_url,
                "filename": filename
            })
        
        # Extract Datawrapper charts
        dw_matches = re.findall(r'datawrapper\.dwcdn\.net/([a-zA-Z0-9]+)/(\d+)', html)
        for chart_id, version in set(dw_matches):
            metadata["datawrapper_charts"].append({
                "chart_id": chart_id,
                "version": version,
                "csv_url": f"https://datawrapper.dwcdn.net/{chart_id}/{version}/dataset.csv"
            })
        
        return metadata
    
    async def search_publications(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        Search SIPRI publications.
        
        Args:
            params: Can contain:
                    - "query": Search term
                    - "year": Publication year
        
        Returns:
            List of publications with URLs
        """
        query = params.get("query", "")
        year = params.get("year", "")
        
        # Build search URL
        search_url = f"{self.BASE_URL}/publications/search"
        if query:
            search_url += f"?search_api_views_fulltext={query.replace(' ', '+')}"
        if year:
            search_url = f"{self.BASE_URL}/publications/{year}"
        
        result = await self._fetch_url(search_url)
        
        if "error" in result:
            return result
        
        html = result["content"].decode('utf-8')
        
        # Extract publication links
        publications = []
        pub_links = re.findall(r'href="(/publications/\d{4}/[^"]+)"', html)
        
        seen = set()
        for link in pub_links:
            if link not in seen:
                seen.add(link)
                # Extract title from URL
                parts = link.split('/')
                if len(parts) >= 4:
                    title_slug = parts[-1]
                    title = title_slug.replace('-', ' ').title()
                    publications.append({
                        "url": f"{self.BASE_URL}{link}",
                        "path": link,
                        "title": title,
                        "year": parts[2]
                    })
        
        return {
            "query": query,
            "year": year,
            "search_url": search_url,
            "publications": publications[:50]  # Limit to 50 results
        }
    
    async def list_databases(self) -> Dict[str, Any]:
        """List all available databases and their descriptions."""
        databases = []
        
        for key, info in self.DATABASE_FILES.items():
            databases.append({
                "key": key,
                "description": info["description"],
                "format": info["format"],
                "url": info["url"],
                "category": info["category"]
            })
        
        charts = []
        for key, info in self.DATAWRAPPER_CHARTS.items():
            charts.append({
                "key": key,
                "chart_id": info["chart_id"],
                "description": info["description"],
                "csv_url": f"https://datawrapper.dwcdn.net/{info['chart_id']}/{info['version']}/dataset.csv"
            })
        
        return {
            "databases": databases,
            "charts": charts
        }
    
    async def close(self):
        """Close the HTTP session."""
        if self.session and not self.session.closed:
            await self.session.close()


async def execute(params: Dict[str, Any], ctx: Any = None) -> Dict[str, Any]:
    """
    Main entry point for the SIPRI access skill.
    
    Args:
        params: Dictionary containing:
                - "function": One of:
                    - "download_database": Download a database file
                    - "get_datawrapper_data": Get data from Datawrapper chart
                    - "get_publication": Get publication info and PDF links
                    - "search_publications": Search for publications
                    - "list_databases": List available databases
                - Additional function-specific parameters
        
        ctx: Optional context (not used)
    
    Returns:
        Dictionary with results or error information
    """
    function = params.get("function")
    
    if not function:
        return {
            "error": "Missing required parameter: function",
            "available_functions": [
                "download_database",
                "get_datawrapper_data",
                "get_publication",
                "search_publications",
                "list_databases"
            ]
        }
    
    skill = SIPRIAccessSkill()
    
    try:
        if function == "download_database":
            return await skill.download_database(params)
        
        elif function == "get_datawrapper_data":
            return await skill.get_datawrapper_data(params)
        
        elif function == "get_publication":
            return await skill.get_publication(params)
        
        elif function == "search_publications":
            return await skill.search_publications(params)
        
        elif function == "list_databases":
            return await skill.list_databases()
        
        else:
            return {
                "error": f"Unknown function: {function}",
                "available_functions": [
                    "download_database",
                    "get_datawrapper_data",
                    "get_publication",
                    "search_publications",
                    "list_databases"
                ]
            }
    
    except Exception as e:
        return {
            "error": f"Execution error: {str(e)}",
            "function": function
        }
    
    finally:
        await skill.close()